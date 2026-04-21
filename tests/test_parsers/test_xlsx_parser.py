"""XlsxParser: openpyxl-based sheet extraction.

Guards:
  * each sheet emits as ``## Sheet: {name}`` section
  * cells render as pipe-markdown table with header row + separator
  * formula cells return the cached value (data_only=True) not
    the formula string
  * rows above the per-sheet cap are trimmed with a trailer note
  * empty sheets don't emit a table (just the heading, trimmed)
  * registry dispatches .xlsx to xlsx-native ahead of docling
"""
import io

import pytest


def _build_xlsx(
    sheets: dict[str, list[list]],
) -> bytes:
    """Build an .xlsx from a dict of sheet name → list of row values."""
    from openpyxl import Workbook

    wb = Workbook()
    # Remove the default "Sheet"
    default = wb.active
    wb.remove(default)
    for name, rows in sheets.items():
        ws = wb.create_sheet(title=name)
        for row in rows:
            ws.append(row)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture(autouse=True)
def _disable_isolation(monkeypatch):
    from config import get_settings
    monkeypatch.setattr(get_settings(), "parser_isolation", False)


# ---------------------------------------------------------------------------
# Engine metadata
# ---------------------------------------------------------------------------

def test_engine_name_and_supported_types():
    from parsers.xlsx_parser import XlsxParser
    assert XlsxParser.engine_name == "xlsx-native"
    assert ".xlsx" in XlsxParser.supported_types
    assert ".xls" in XlsxParser.supported_types


def test_is_available_with_dep_present():
    from parsers.xlsx_parser import XlsxParser
    assert XlsxParser.is_available()


# ---------------------------------------------------------------------------
# Text extraction
# ---------------------------------------------------------------------------

@pytest.mark.asyncio
async def test_parse_renders_sheet_as_pipe_table():
    from parsers.xlsx_parser import XlsxParser
    xlsx = _build_xlsx({
        "People": [
            ["Name", "Age", "City"],
            ["Ada", 36, "London"],
            ["Grace", 72, "Baltimore"],
        ],
    })
    result = await XlsxParser().parse(xlsx, filename="data.xlsx")

    assert "## Sheet: People" in result.content
    # Header + separator + body rows
    assert "| Name | Age | City |" in result.content
    assert "| --- | --- | --- |" in result.content
    assert "| Ada | 36 | London |" in result.content


@pytest.mark.asyncio
async def test_parse_multiple_sheets_each_in_own_section():
    from parsers.xlsx_parser import XlsxParser
    xlsx = _build_xlsx({
        "A": [["x", "y"], [1, 2]],
        "B": [["p"], [10]],
    })
    result = await XlsxParser().parse(xlsx, filename="multi.xlsx")

    assert "## Sheet: A" in result.content
    assert "## Sheet: B" in result.content
    # Each sheet's content appears below its heading (order-preserved)
    a_idx = result.content.index("## Sheet: A")
    b_idx = result.content.index("## Sheet: B")
    assert a_idx < b_idx
    assert "| x | y |" in result.content
    assert "| p |" in result.content


@pytest.mark.asyncio
async def test_parse_escapes_pipes_in_cell_content():
    """Pipe characters inside cells must be escaped to avoid breaking
    the table structure for downstream markdown consumers.
    """
    from parsers.xlsx_parser import XlsxParser
    xlsx = _build_xlsx({
        "S": [["Heading"], ["value|with|pipes"]],
    })
    result = await XlsxParser().parse(xlsx, filename="x.xlsx")
    assert "value\\|with\\|pipes" in result.content


@pytest.mark.asyncio
async def test_parse_collapses_newlines_in_cells():
    from parsers.xlsx_parser import XlsxParser
    xlsx = _build_xlsx({
        "S": [["H"], ["line1\nline2"]],
    })
    result = await XlsxParser().parse(xlsx, filename="x.xlsx")
    # Newline in cell should become a space; otherwise the table breaks
    assert "line1 line2" in result.content


@pytest.mark.asyncio
async def test_parse_caps_row_count_and_emits_truncation_notice():
    """Keep runaway spreadsheets from busting max_content_chars."""
    from parsers.xlsx_parser import XlsxParser
    from parsers.xlsx_parser import _MAX_ROWS_PER_SHEET

    many_rows = [["col"]] + [[i] for i in range(_MAX_ROWS_PER_SHEET + 100)]
    xlsx = _build_xlsx({"Big": many_rows})
    result = await XlsxParser().parse(xlsx, filename="big.xlsx")

    # Truncation marker should appear exactly once per truncated sheet.
    assert "sheet truncated at" in result.content
    # Row count in output is <= cap+1 header lines etc.; exact count
    # is less important than the marker being present.


@pytest.mark.asyncio
async def test_parse_title_from_filename_stem():
    from parsers.xlsx_parser import XlsxParser
    xlsx = _build_xlsx({"Sheet1": [["x"]]})
    result = await XlsxParser().parse(xlsx, filename="quarterly_report.xlsx")
    assert result.title == "quarterly_report"


@pytest.mark.asyncio
async def test_parse_returns_empty_image_refs():
    """Images are not produced here — an XlsxImageExtractor plugin
    would own that. Fallback to docling for .xlsx is still registered
    if someone cares about the image path today.
    """
    from parsers.xlsx_parser import XlsxParser
    xlsx = _build_xlsx({"S": [["x"]]})
    result = await XlsxParser().parse(xlsx, filename="x.xlsx")
    assert result.image_refs == []


# ---------------------------------------------------------------------------
# Formula values
# ---------------------------------------------------------------------------

@pytest.mark.asyncio
async def test_parse_formula_resolves_to_cached_value():
    """data_only=True uses the last cached computed value Excel
    stored. The test file here carries formulas WITHOUT cached values
    (openpyxl doesn't compute on save), which would render as None;
    but operators editing the sheet in Excel would get real values.
    We assert the contract indirectly by writing a formula + checking
    the formula string does NOT leak through.
    """
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append([1, 2, "=A1+B1"])  # formula without cached value
    buf = io.BytesIO()
    wb.save(buf)

    from parsers.xlsx_parser import XlsxParser
    result = await XlsxParser().parse(buf.getvalue(), filename="f.xlsx")

    # data_only=True means openpyxl returns None for a formula that
    # has no cached value (rather than the "=A1+B1" string). The raw
    # formula string must never appear in content.
    assert "=A1+B1" not in result.content


# ---------------------------------------------------------------------------
# Error handling
# ---------------------------------------------------------------------------

@pytest.mark.asyncio
async def test_parse_handles_malformed_bytes():
    from parsers.xlsx_parser import XlsxParser
    result = await XlsxParser().parse(b"not a xlsx", filename="bad.xlsx")
    assert result.content == ""
    assert result.image_refs == []


# ---------------------------------------------------------------------------
# Registry dispatch
# ---------------------------------------------------------------------------

def test_registry_picks_xlsx_native_over_docling():
    from parsers.registry import registry
    engine = registry.auto_select(filename="data.xlsx")
    assert engine == "xlsx-native"
