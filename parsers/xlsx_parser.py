"""Native .xlsx parser via openpyxl.

Why native vs docling
---------------------
docling walks .xlsx via openpyxl internally, same as we do. Going
direct avoids the convert_lock (PDF-thread-safety guard that .xlsx
doesn't need) and skips the ~400 MB model-load path entirely on
workers that only ever see spreadsheets.

Output shape
------------
Each sheet becomes a ``## Sheet: {name}`` section. Rows render as
pipe-markdown tables, bounded by ``_MAX_ROWS_PER_SHEET`` /
``_MAX_COLS`` to keep a tab with a million rows from blowing past
``max_content_chars``. When a sheet is truncated, a trailer line
records the skip so RAG retrieval surfaces "there's more data".

Memory
------
``load_workbook(read_only=True, data_only=True)``:
  * read_only streams cells instead of materialising the whole
    workbook in an object tree — bounded memory on 100 MB+ sheets
  * data_only returns cached formula results (``=SUM(A1:A100)`` →
    ``450``) which is what we want for downstream embedding; raw
    formulas would just confuse the LLM

Images are not produced here — would belong to a future
XlsxImageExtractor plugin (W4-b framework). For now the registry's
docling fallback still handles them if anyone cares.
"""
from __future__ import annotations

import asyncio
import io
import logging
from pathlib import Path

from config import get_settings
from parsers.base import BaseParser, ParseResult
from parsers.convert import LEGACY_FORMAT_MAP, convert_legacy_format
from parsers.isolation import run_isolated

logger = logging.getLogger(__name__)

# Hard caps to keep runaway spreadsheets from piling into
# ``max_content_chars``. Tune if operators complain about truncated
# RAG content; most business / analytical spreadsheets fit well under
# these limits.
_MAX_ROWS_PER_SHEET = 500
_MAX_COLS = 50


class XlsxParser(BaseParser):
    engine_name = "xlsx-native"
    supported_types = [".xlsx", ".xls"]

    @classmethod
    def is_available(cls) -> bool:
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            return False
        return True

    async def parse(
        self, source: bytes | str, filename: str = "", **kwargs,
    ) -> ParseResult:
        cfg = get_settings()
        if not cfg.parser_isolation:
            return await asyncio.get_running_loop().run_in_executor(
                None, self._parse_sync, source, filename,
            )
        return await run_isolated(
            _xlsx_parse_worker, source, filename, timeout=cfg.parser_timeout,
        )

    def _parse_sync(self, source: bytes | str, filename: str) -> ParseResult:
        import openpyxl

        raw = source if isinstance(source, bytes) else source.encode()

        suffix = Path(filename).suffix.lower()
        if suffix in LEGACY_FORMAT_MAP:
            import tempfile
            with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
                tmp.write(raw)
                tmp_path = tmp.name
            try:
                converted_path = convert_legacy_format(tmp_path, suffix)
                try:
                    with open(converted_path, "rb") as f:
                        raw = f.read()
                finally:
                    Path(converted_path).unlink(missing_ok=True)
            finally:
                Path(tmp_path).unlink(missing_ok=True)

        try:
            wb = openpyxl.load_workbook(
                io.BytesIO(raw), read_only=True, data_only=True,
            )
        except Exception:
            logger.exception("openpyxl load failed for %s", filename)
            return ParseResult(content="", title=Path(filename).stem)

        lines: list[str] = []
        title = Path(filename).stem
        lines.append(f"# {title}")
        lines.append("")

        try:
            for sheet in wb.worksheets:
                lines.append(f"## Sheet: {sheet.title}")
                lines.append("")
                sheet_md, truncated = _sheet_to_markdown(sheet)
                if sheet_md:
                    lines.append(sheet_md)
                    if truncated:
                        lines.append("")
                        lines.append(
                            f"_(sheet truncated at {_MAX_ROWS_PER_SHEET} rows "
                            f"× {_MAX_COLS} columns; further data omitted)_"
                        )
                    lines.append("")
        finally:
            # read_only workbooks hold file handles on zip entries —
            # close them explicitly to prevent tmp-file leaks under
            # high ingest throughput.
            try:
                wb.close()
            except Exception:
                pass

        return ParseResult(
            content="\n".join(lines).strip() + "\n",
            title=title,
            image_refs=[],
        )


def _sheet_to_markdown(sheet) -> tuple[str, bool]:
    """Render the sheet as a pipe-markdown table up to the row/col cap.

    Returns ``(markdown, truncated)``. Empty sheet returns ``("",
    False)``. Leading empty rows are skipped; interior empty rows are
    preserved as blank cells so column alignment stays consistent.
    """
    rows_out: list[list[str]] = []
    truncated = False
    max_cols = 0
    empty_seen = 0

    for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
        if r_idx >= _MAX_ROWS_PER_SHEET:
            truncated = True
            break
        cells = [_cell_str(v) for v in row[:_MAX_COLS]]
        if len(row) > _MAX_COLS:
            truncated = True
        # Skip leading empty rows; keep interior ones once any content
        # has appeared.
        is_empty = not any(c for c in cells)
        if is_empty and not rows_out:
            empty_seen += 1
            continue
        rows_out.append(cells)
        max_cols = max(max_cols, len(cells))

    # Strip trailing empty rows we may have accumulated.
    while rows_out and not any(rows_out[-1]):
        rows_out.pop()

    if not rows_out:
        return "", False

    # Pad rows to uniform column count so the pipe-table is rectangular.
    for row in rows_out:
        if len(row) < max_cols:
            row.extend([""] * (max_cols - len(row)))

    # Header separator: first row treated as header.
    header = "| " + " | ".join(rows_out[0]) + " |"
    sep = "| " + " | ".join(["---"] * max_cols) + " |"
    body = "\n".join(
        "| " + " | ".join(r) + " |" for r in rows_out[1:]
    )
    if body:
        return f"{header}\n{sep}\n{body}", truncated
    return f"{header}\n{sep}", truncated


def _cell_str(value) -> str:
    """Coerce a cell value to a markdown-safe string.

    Numbers and dates: ``str(value)`` (openpyxl returns datetime
    objects for date-formatted cells; default str form is fine for
    RAG embedding). ``None`` → empty. Pipes and newlines are escaped
    so the table doesn't break.
    """
    if value is None:
        return ""
    return str(value).replace("|", "\\|").replace("\n", " ").strip()


def _xlsx_parse_worker(source: bytes | str, filename: str) -> ParseResult:
    """Subprocess entry for XlsxParser."""
    return XlsxParser()._parse_sync(source, filename)
